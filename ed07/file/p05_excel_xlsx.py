'''
https://docs.python.org/ko/3/reference
'''

from openpyxl import load_workbook

def main():
    # 엑셀 파일 읽기
    wb = load_workbook('KOBIS_주간_주말_박스오피스_2023-07-19.xlsx')  # workbook
    ws = wb.active  # 첫 번째 sheet

    for row in ws.iter_rows(min_row=9, values_only=True):
        # print('row', type(row))  # row <class 'tuple'>
        순위 = row[0]
        영화명 = row[1]
        점유율 = round(row[4] * 100, 2)
        관객수 = row[8]

        print(f'순위: {순위}, 영화명: {영화명}, 점유율: {점유율}%, 관객수: {관객수}명')

    wb.close()

main()
